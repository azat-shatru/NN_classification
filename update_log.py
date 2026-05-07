import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.load_workbook('NN_Design_Tracker.xlsx')
ws = wb['Session Log']

next_row = ws.max_row + 1
next_session = 24

decision_text = (
    "Session 2026-05-07 — Charts Portal: drag-and-drop card reordering + slide reordering. "
    "assets/chart_card_dnd.js: HTML5 drag-and-drop for .cp2c-chart-card elements inside "
    "#cp2c-cards-container; uses mid-point x detection to insert before/after target in flex grid. "
    "assets/slide_order_dnd.js: HTML5 drag-and-drop for .slide-order-row elements in the Export card; "
    "flashes purple outline on container after drop to hint user to click Confirm. "
    "stage2c_charts.py: cards now rendered as flat flex container (replacing nested dbc.Row/Col); "
    "card width computed dynamically from cols_per_row slider as calc(N% - Xpx); "
    "each card wrapper has draggable=true, data-code attr, and a braille-dot drag handle in the header. "
    "Save Order button (top-right, next to Apply): clientside callback reads DOM .cp2c-chart-card order "
    "and saves list of codes to cp2c-card-order session store. "
    "render_charts sorts visible tiles by stored card order before rendering; unordered tiles appended at end. "
    "Slide Order panel in Export card: each slide group shown as a draggable row with chart code chips. "
    "Confirm Order button: clientside callback reads DOM .slide-order-row order and saves to cp2c-slide-order. "
    "build_pptx() in ppt_export.py now accepts slide_order list; iterates slides in confirmed order; "
    "phantom slide numbers silently skipped; new slides appended at end. "
    "Card height fix: container uses align-items:stretch; card wrappers are display:flex + flex-direction:column; "
    "dbc.Card has height:100% so all cards in a row stretch to match the tallest — prevents size "
    "inconsistency after drag-and-drop reorder. Removed duplicate margin-bottom (gap handles row spacing). "
    "render_charts callback: 3 outputs (cp2c-charts-grid, cp2c-summary, cp2c-slide-rows); "
    "states include both cp2c-card-order and cp2c-slide-order. "
    "Tested with real data: 1,200 rows x 713 cols, 54 questions, 49 codebook tiles — all tests passed."
)

open_questions = (
    "A7 value=8 (NEE) custom missing value handling still pending (Stage 1). "
    "Test full pipeline end-to-end with Raw data 2.xlsx + Survey 2.docx. "
    "Re-zip and upload to Google Drive. "
    "Stages 1-9 inherited from Streamlit rewrite, not yet re-tested end-to-end. "
    "S10B root cause (page-break split table vs vMerge skip on rows 10-20) still open."
)


# Copy style from previous data row
def copy_row_style(src_row_num, dst_row_num):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row_num, column=col)
        dst_cell = ws.cell(row=dst_row_num, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

copy_row_style(ws.max_row, next_row)

ws.cell(row=next_row, column=1).value = next_session
ws.cell(row=next_row, column=2).value = '2026-05-07 S24'
ws.cell(row=next_row, column=3).value = 'Charts Portal: drag-and-drop card + slide reorder, card height fix'
ws.cell(row=next_row, column=4).value = decision_text
ws.cell(row=next_row, column=5).value = open_questions

ws.row_dimensions[next_row].height = 200

wb.save('NN_Design_Tracker.xlsx')
print(f"Saved. Session {next_session} added at row {next_row}.")
