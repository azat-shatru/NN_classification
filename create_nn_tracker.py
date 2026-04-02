import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Colour palette ──────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
YELLOW      = "FFF2CC"
GREEN       = "E2EFDA"
ORANGE      = "FCE4D6"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(color="000000", size=11):
    return Font(bold=True, color=color, size=size)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ════════════════════════════════════════════════════════════════
# SHEET 1 – Design Decisions
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Design Decisions"

# ── Title banner ────────────────────────────────────────────────
ws1.merge_cells("A1:G1")
ws1["A1"] = "Neural Network Design Tracker — Multiclass Classification"
ws1["A1"].font      = Font(bold=True, color=WHITE, size=14)
ws1["A1"].fill      = fill(DARK_BLUE)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:G2")
ws1["A2"] = "Dataset size (N): 300–600    |    Task: Multiclass Classification (4 classes)    |    Last updated: 2026-04-03    |    ALL 11 STAGES COMPLETE + UI POLISH"
ws1["A2"].font      = Font(italic=True, color=WHITE, size=10)
ws1["A2"].fill      = fill(MID_BLUE)
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

# ── Column headers ───────────────────────────────────────────────
headers = ["Category", "Parameter", "Options / Range", "Recommendation\n(small N)", "Your Choice", "Status", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font      = bold(WHITE, 10)
    c.fill      = fill(MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
ws1.row_dimensions[3].height = 30

# ── Parameter rows ───────────────────────────────────────────────
rows = [
    # Category, Parameter, Options/Range, Recommendation, Your Choice, Status, Notes
    ("Problem Setup",   "Number of input features",   "Any positive integer",              "Start 50–100; reduce to 5–10 via feature selection", "50–100 raw → target 5–10", "In Progress", "Mixed: numeric + categorical. Full pipeline below."),
    ("Problem Setup",   "Number of classes",          "Any integer ≥ 2",                   "—",                              "4",   "Done",       ""),
    ("Problem Setup",   "Class balance",              "Balanced / Imbalanced",             "Check before training",          "Balanced (≤20% variation across classes)", "Done", "Accuracy is valid metric; no SMOTE needed"),
    ("Problem Setup",   "Input variable types",       "Numeric / Categorical / Mixed",     "Handle separately in pipeline",  "Mixed (numeric + categorical)", "Done", "Categorical needs encoding before scaling/NN"),

    ("Data Quality",    "Stage 0: Data Audit",        "Shape, dtypes, missing %, duplicates, stats, cardinality", "Always run first; visual dashboard", "Visual dashboard: missing heatmap, type table, class dist., cardinality", "Done", "stage0_audit.py — upload CSV/Excel, column type assignment, missing heatmap, class dist bar, numeric stats table"),
    ("Data Quality",    "Missing value threshold",    "Drop if > X% missing",              "Drop if >40% missing per variable", "User-adjustable slider (default 40%)", "Done", "Interactive multiselect to confirm columns to drop"),
    ("Data Quality",    "Missing value imputation — numeric",   "Mean / Median / KNN",    "Median (robust to outliers); KNN for structured missingness", "Median / Mean / KNN (user selects)", "Done", "stage1_missing.py — fit on train only, transform both"),
    ("Data Quality",    "Missing value imputation — categorical", "Mode / 'Unknown' category", "Add 'Unknown' category — preserves missingness signal", "Mode / Unknown constant (user selects)", "Done", ""),
    ("Data Quality",    "Duplicate row handling",     "Drop / Review",                     "Drop exact duplicates; flag near-duplicates for review", "Duplicate count shown in Stage 0 audit", "Done", ""),
    ("Data Quality",    "Train/test split timing",    "Before or after preprocessing",     "ALWAYS split before imputation/scaling to avoid data leakage", "Split first, then preprocess", "Done", "Critical: fit imputer/scaler on train only, transform both"),

    ("Outlier Handling","Stage 2: Outlier detection method", "IQR / Z-score / Isolation Forest", "IQR boxplot per variable (visual) + Isolation Forest (multivariate)", "IQR boxplot + Isolation Forest", "Done", "stage2_outliers.py — per-column boxplots, IQR/Z-score detection, Isolation Forest multivariate"),
    ("Outlier Handling","Outlier treatment options",  "Cap/Winsorize / Remove / Transform / Keep", "Winsorize (1st–99th pct) as default; user overrides per variable", "Winsorize / Remove rows / Log-transform / Keep (per column)", "Done", "User picks treatment per column; preview before/after distribution"),
    ("Outlier Handling","Skewness correction",        "Log / Box-Cox / Yeo-Johnson / None","Yeo-Johnson (handles negatives); apply to skewed numeric vars", "Log-transform option available per column", "Done", ""),

    ("Encoding",        "Stage 3a: Nominal encoding (low cardinality ≤10)", "One-Hot Encoding", "One-Hot; drop first to avoid multicollinearity", "One-Hot Encoding — pd.get_dummies, test aligned to train columns", "Done", "stage3_encoding.py — auto-detects binary/cat/ordinal; user can override per column"),
    ("Encoding",        "Stage 3b: Nominal encoding (high cardinality >10)", "Target/Mean Encoding / Hashing", "Target encoding (use with CV to avoid leakage)", "Target Encoding — mean(target) per category, fitted on train only", "Done", "Unseen test categories filled with global mean"),
    ("Encoding",        "Stage 3c: Ordinal encoding", "Ordinal / Label",                  "Ordinal encoding preserving rank order",          "Ordinal Encoding — user defines category order per column in text area", "Done", "User types one category per line (lowest → highest); fallback = len(order)"),
    ("Encoding",        "Stage 3d: Binary variables", "Label Encoding (0/1)",             "Label encode directly",                          "Binary (keep as-is) for ≤2 unique values; Label Encoding also available", "Done", "Auto-detected and skipped from encoding"),

    ("Scaling",         "Stage 4: Numeric scaling",   "StandardScaler / MinMaxScaler / RobustScaler", "RobustScaler if outliers present; StandardScaler otherwise", "RobustScaler default; per-column override + exclude list available", "Done", "stage4_scaling.py — binary cols auto-skipped; scalers fitted on train only"),
    ("Scaling",         "Train/test split enforcement", "Split BEFORE imputation and scaling", "CRITICAL: fit imputer+scaler on train only; transform test separately", "Split first → fit on train → transform both", "Done", "Prevents data leakage — test set must never influence preprocessing"),

    ("Visualization",   "Stage 2.5: Viz Dashboard placement", "After outlier treatment; before encoding/scaling", "Show cleaned but un-transformed data with readable labels", "Post-clean, pre-encoding, pre-scaling", "Done", "Charts show real distributions before any numeric transformation"),
    ("Visualization",   "App framework",              "Streamlit / Dash / Jupyter widgets",  "Streamlit — easiest to build; runs locally in browser", "Streamlit", "Done", "app/app.py — 11-stage sidebar navigation with ✅/⬜ status + progress bar"),
    ("Visualization",   "Codebook input method",      "QNR doc upload / manual entry / auto-map", "QNR doc auto-parse + manual edit + dataset auto-map", "Auto-parse from QNR doc → build_codebook() → editable per-question expanders", "Done", "Supports .docx/.pdf/.xlsx/.txt questionnaire upload"),
    ("Visualization",   "Variable types supported",   "Categorical(0/1) / Ordinal / Numeric / scale_7 / scale_5 / multi / grid / open", "All types — different chart options per type", "9 variable types with type-specific chart menus", "Done", "Type drives chart_options; user can override per tile"),
    ("Visualization",   "Value → label mapping",      "Per-variable, per-value text labels",  "Categorical: 0/1 → 2 labels. Ordinal: each int → label. Numeric: none", "st.data_editor table per question: option_label, dataset_col, value_coding, notes", "Done", "Dynamic rows (+ button to add); dataset_col is SelectboxColumn over all df columns"),
    ("Visualization",   "Drag-and-drop interface",    "streamlit-sortables component",        "Drag variables to question/option slots; fallback to multiselect", "st.data_editor with SelectboxColumn for column assignment; search + filter toolbar", "Done", "Add/delete rows dynamically; all_cols synced back from options table"),
    ("Visualization",   "Chart types — Categorical",  "Bar chart / Pie chart / Donut chart",  "Bar chart (counts or %)",                "User selects per variable", "Done", ""),
    ("Visualization",   "Chart types — Ordinal",      "Bar chart / Stacked bar / Line / Horizontal bar", "Bar chart with ordered labels on x-axis","User selects per variable", "Done", ""),
    ("Visualization",   "Chart types — Numeric",      "Histogram / Box plot / Violin / Line", "Histogram (distribution) + Box plot (outliers)", "User selects per variable", "Done", ""),
    ("Visualization",   "Chart export",               "PNG / PDF / HTML",                     "Export all charts as PDF report",        "PDF export via reportlab (Step 5)", "Done", "Requires kaleido for Plotly PNG export; pip install kaleido"),
    ("Visualization",   "Questionnaire doc upload",   "PDF / Word (.docx) / Excel / TXT",     "Auto-parse to codebook draft; user edits misparses", "Upload + auto-parse + editable expanders per question", "Done", "utils/qnr_parser.py — regex extracts codes, text, type, scale anchors, options"),
    ("Visualization",   "Parsing targets",            "Question numbers, text, options, scales","Regex: '1.', 'Q1.', 'a)', bullets, Likert labels", "Q code, question text, q_type, scale_low/high, scale_points, options, pn_notes", "Done", "Tested: 31 questions from survey.docx; 29 matched to MyRawdata.xlsx columns"),
    ("Visualization",   "Tile edit on misparse",      "Full manual editor — add, edit, delete, reassign", "Each question is an expander; all fields editable including options table", "Per-question expanders: code, text, type, chart, column, options table, value labels", "Done", "+ Add new question form; Delete button per question; search + filter toolbar"),
    ("Visualization",   "Dashboard layout mode",      "Grid / Slide",                          "Both modes; user toggles",               "Grid mode + Slide mode with prev/next navigation", "Done", ""),
    ("Visualization",   "Tiles per row (grid mode)",  "1 / 2 / 3 / 4",                        "3 tiles per row default",                "User selects: 1–4 per row", "Done", ""),
    ("Visualization",   "Tiles per slide (slide mode)","1 – 12",                               "6 tiles per slide default",              "User selects; prev/next navigation", "Done", "Slide mode for presentations; grid mode for analysis"),

    ("Feature Selection", "Stage 5: Correlation Analysis", "Pearson / Spearman",          "Pearson for continuous; visual heatmap", "Interactive Plotly heatmap; threshold slider; flagged pair table; scatter drill-down; user selects drops", "Done", "stage5_correlation.py — suggested drops = 2nd col of each correlated pair; user can override"),
    ("Feature Selection", "Stage 6: RF Importance",   "Random Forest feature importances", "Bar chart; user selects features to keep", "RF importance bar chart + cumulative curve; 3 selection modes: Top-N, cumulative %, manual", "Done", "stage6_rf.py — configurable trees/depth/seed; apply removes columns from X_train+X_test"),
    ("Feature Selection", "Stage 7: Factor Analysis", "EFA with varimax / oblique rotation","Loadings table + biplot; user names or picks high-loaders", "Bartlett + KMO tests; scree plot; loadings table with highlighting; biplot; 3 output modes", "Done", "stage7_factor.py — output as factor scores, high-loaders, or both; skip option available"),
    ("Feature Selection", "Target feature count",     "Any positive integer",              "5–10 goal; may be 10–25 after Stage 3", "5–10 goal (flexible)",  "Done", "Hard target relaxed — best subset found via Stage 8"),
    ("Feature Selection", "Stage 8: Combination Testing — automated", "Forward stepwise / Backward stepwise / Exhaustive subset", "Forward + Backward stepwise scored by CV accuracy + AUC", "Forward + Backward stepwise with configurable metric and CV folds", "Done", "stage8_combtest.py — proxy model: LogisticRegression; scores: Accuracy, AUC, McFadden R²"),
    ("Feature Selection", "Stage 8: Combination Testing — manual",    "User-defined subset explorer",       "Dashboard: user picks subset, model trains, all metrics shown", "Manual explorer tab: pick any combination, score instantly, set for Stage 9", "Done", "Results leaderboard across all scored combinations; forward/backward/manual results compared"),
    ("Feature Selection", "Stage 8: Metrics per combination",         "Accuracy / AUC / Confusion Matrix / SHAP / Pseudo-R²", "All 5 metrics per combination", "Accuracy + AUC (macro) + Confusion Matrix + SHAP + McFadden R²", "Done", "SHAP replaces Wald/p-value for NNs; pseudo-R² replaces R-squared"),
    ("Architecture",    "Framework",                  "PyTorch",                           "PyTorch — confirmed by user",    "PyTorch (nn.Module)",  "Done", "Custom training loop; DataLoader; nn.Module subclass; shap.DeepExplainer for SHAP"),
    ("Architecture",    "PyTorch key components",     "nn.Module / DataLoader / Dataset",  "Subclass nn.Module; use DataLoader for batching; custom train loop", "nn.Module + DataLoader + custom loop", "Done", "TorchTabular or custom Dataset class for tabular data"),
    ("Architecture",    "Input layer size",           "5 – 10 (set at runtime)",           "Passed as parameter — not hardcoded", "input_size = n_selected_features (5–10, set after Stage 8)", "Done", "TabularNN(input_size=...) called after feature selection completes"),
    ("Architecture",    "Number of hidden layers",    "2",                                 "2 for small N",                  "2",  "Done", ""),
    ("Architecture",    "Hidden layer 1 neurons",     "2–4 × input_size",                  "16–32 depending on input size",  "hidden1 = 2–4 × input_size (e.g. 16–32)", "Done", "Scales with selected feature count"),
    ("Architecture",    "Hidden layer 2 neurons",     "hidden1 / 2",                       "8–16",                           "hidden2 = hidden1 // 2 (e.g. 8–16)", "Done", "Funnel shape"),
    ("Architecture",    "Layer shape",                "Funnel",                            "Funnel: input→hidden1→hidden2→4", "Funnel: input_size → hidden1 → hidden2 → 4", "Done", ""),
    ("Architecture",    "Output layer",               "nn.Linear(hidden2, 4)",             "4 neurons = 4 classes; raw logits", "nn.Linear(hidden2, 4) — raw logits", "Done", ""),
    ("Architecture",    "Output activation",          "None in forward() — handled by loss","CrossEntropyLoss includes softmax", "No Softmax in forward()", "Done", "PyTorch CrossEntropyLoss = log_softmax + NLLLoss internally"),
    ("Activation",      "Hidden layer activation",    "nn.ReLU",                           "nn.ReLU",                        "nn.ReLU", "Done", "He init (PyTorch default for Linear) pairs well with ReLU"),
    ("Regularization",  "Dropout rate",               "0.3",                               "0.3 after each hidden layer",    "0.3",  "Done", "Applied after ReLU, before next Linear layer"),
    ("Regularization",  "L2 weight decay",            "1e-4",                              "1e-4 in Adam weight_decay param","1e-4 via Adam(weight_decay=1e-4)", "Done", ""),
    ("Regularization",  "Batch Normalization",        "No (start without)",                "Try without first; add if unstable","No — start without", "Done", ""),
    ("Loss",            "Loss function",              "nn.CrossEntropyLoss",               "nn.CrossEntropyLoss",            "nn.CrossEntropyLoss", "Done", "Expects raw logits; targets as class indices (0–3)"),
    ("Optimizer",       "Optimizer",                  "torch.optim.Adam",                  "torch.optim.Adam",               "torch.optim.Adam", "Done", ""),
    ("Optimizer",       "Learning rate",              "1e-3",                              "1e-3 (Adam default)",            "1e-3", "Done", ""),
    ("Optimizer",       "Weight decay",               "1e-4",                              "1e-4",                           "1e-4", "Done", "Passed as Adam(weight_decay=1e-4)"),
    ("Optimizer",       "LR scheduling",              "ReduceLROnPlateau",                 "ReduceLROnPlateau(patience=10)",  "ReduceLROnPlateau(patience=10, factor=0.5)", "Done", ""),
    ("Training",        "Batch size",                 "32",                                "32 for N=300–600",               "32",  "Done", ""),
    ("Training",        "Max epochs",                 "300",                               "300 + early stopping",           "300", "Done", ""),
    ("Training",        "Early stopping patience",    "15",                                "15 epochs",                      "15", "Done", "Monitor val loss; restore best weights"),
    ("Training",        "Validation strategy",        "k-Fold CV (k=5)",                   "k-Fold strongly recommended for small N", "5-Fold CV", "Done", ""),
    ("Training",        "Train/val split",            "80/20 within each fold",            "80/20",                          "80/20 per fold", "Done", ""),
    ("Initialization",  "Weight initialization",      "He (Kaiming)",                      "PyTorch default for nn.Linear — He uniform", "He uniform (PyTorch default)", "Done", "No manual init needed — PyTorch applies this automatically"),
    ("Imbalance",       "Class imbalance handling",   "Class weights / SMOTE / None",      "Class weights if imbalanced",    "None (balanced dataset)",  "Done", "Max 20% variation — no intervention needed"),
    ("Evaluation",      "Primary metric",             "Accuracy / F1 / AUC-ROC",          "Macro-F1 if imbalanced",         "Accuracy",  "Done", "Balanced dataset makes accuracy a valid metric"),
    ("Evaluation",      "Confusion matrix",           "Yes / No",                          "Yes — always inspect",           "Yes", "Done", "Per-class breakdown essential for 4-class problem"),
    ("Evaluation",      "AUC-ROC",                    "Per-class (one-vs-rest) + macro avg","Yes — macro AUC for overall model",  "Macro AUC + per-class AUC", "Done", "4 classes: compute one-vs-rest AUC for each, then average"),
    ("Evaluation",      "SHAP values",                "Per-feature, per-class contribution","Yes — replaces Wald/p-value for NN", "SHAP summary + beeswarm plots", "Done", "Shows direction + magnitude of each variable's contribution"),
    ("Evaluation",      "McFadden pseudo-R²",         "0 to 1 (>0.2 good, >0.4 excellent)", "Yes — classification R² equivalent", "McFadden pseudo-R²", "Done", "Computed from log-likelihood; classification equivalent of R-squared"),
]

category_colors = {
    "Problem Setup":    LIGHT_BLUE,
    "Data Quality":     "FCE4D6",
    "Outlier Handling": "FFF2CC",
    "Encoding":         "E2EFDA",
    "Scaling":          "D6E4F0",
    "Visualization":    "EAD1DC",
    "Feature Selection": "D9EAD3",
    "Architecture":     YELLOW,
    "Activation":     GREEN,
    "Loss":           ORANGE,
    "Optimizer":      LIGHT_BLUE,
    "Regularization": YELLOW,
    "Training":       GREEN,
    "Initialization": ORANGE,
    "Imbalance":      LIGHT_BLUE,
    "Evaluation":     YELLOW,
}

for i, row_data in enumerate(rows):
    r = i + 4
    row_fill = fill(category_colors.get(row_data[0], WHITE))
    for col, val in enumerate(row_data, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.border    = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if col in (1, 2, 3, 4, 8):
            c.fill = row_fill
        elif col == 5:
            c.fill = fill(WHITE)
        elif col == 6:
            c.fill = fill(GREY)
    ws1.row_dimensions[r].height = 18

# ── Column widths ────────────────────────────────────────────────
widths = [16, 26, 32, 34, 22, 12, 30]
for col, w in enumerate(widths, 1):
    set_col_width(ws1, col, w)

# ── Freeze panes ─────────────────────────────────────────────────
ws1.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════
# SHEET 2 – Session Log
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Session Log")

ws2.merge_cells("A1:E1")
ws2["A1"] = "Session Log — paste key decisions / Q&A here to resume in a new Claude instance"
ws2["A1"].font      = Font(bold=True, color=WHITE, size=12)
ws2["A1"].fill      = fill(DARK_BLUE)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

log_headers = ["Session #", "Date", "Topic / Parameter", "Decision / Response", "Open Questions"]
for col, h in enumerate(log_headers, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font      = bold(WHITE, 10)
    c.fill      = fill(MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()

# Pre-fill sessions
session1 = [
    (1, "2026-04-02",
     "Initial design discussion",
     "N=300–600. Discussed all design parameters: architecture, activation, loss, optimizer, regularization, training strategy, evaluation.",
     "How many features? How many classes? Is data balanced?"),
    (2, "2026-04-02",
     "Problem setup + feature selection",
     "4 output classes confirmed. Raw input: 50–100 variables. Goal: reduce to 5–10 best features. Feature selection pipeline agreed: correlation filter → RF importance → RFE. Explained balanced vs imbalanced datasets.",
     "How many samples per class? (needed to decide class imbalance handling and evaluation metric)"),
    (3, "2026-04-02",
     "Class balance + full feature selection pipeline",
     "Dataset confirmed balanced (≤20% variation). Accuracy chosen as primary metric. No SMOTE needed. 3-stage interactive feature selection pipeline agreed: (1) Correlation heatmap — user manually removes pairs, (2) RF importance bar chart — user manually selects, (3) Factor Analysis — user names factor or picks high-loading raw variables. Learning resources recommended: StatQuest (YouTube), Kaggle Learn Feature Engineering, Coursera Feature Engineering course.",
     "Ready to build: which stage do you want to code first? Also still need architecture decisions: hidden layers, neurons, activation, loss, optimizer, batch size."),
    (4, "2026-04-02",
     "Stage 4 feature combination testing + evaluation metrics",
     "Feature target relaxed: 5–10 is a goal not a hard limit. Stage 4 added: automated (forward/backward stepwise) + manual (interactive explorer) combination testing of remaining features. Full metrics dashboard per combination: Accuracy, Macro AUC-ROC (one-vs-rest per class), Confusion Matrix, SHAP values (replaces Wald/p-value for NNs), McFadden pseudo-R² (replaces R-squared for classification). Clarified: Wald/p-values are logistic regression concepts; for NNs use SHAP instead. Learning resources recommended: StatQuest (ROC/AUC, SHAP, Confusion Matrix), Andrew Ng ML Specialization (Coursera), ISLR book Ch.6 (free), Penn State STAT 504 (Wald/logistic regression), fast.ai.",
     "Still pending: architecture parameters (hidden layers, neurons, activation, loss, optimizer, batch size). Then start building pipeline code."),
    (5, "2026-04-02",
     "Data quality + preprocessing pipeline added",
     "Full pipeline now has 8 stages before NN training. New stages added: Stage 0 (Data Audit: visual dashboard of shape, types, missing %, duplicates, class distribution, cardinality), Stage 1 (Missing values: threshold drop + imputation — median/KNN for numeric, mode/Unknown for categorical), Stage 2 (Outlier detection: IQR boxplots + Isolation Forest, treatment: winsorize/remove/transform/keep), Stage 3 (Encoding: One-Hot for low-card nominal, Target encoding for high-card, Ordinal for ranked, Label for binary), Stage 4 (Scaling: RobustScaler if outliers, StandardScaler otherwise — numeric only). Key rule: train/test split MUST happen before any imputation or scaling to avoid data leakage. Input confirmed as mixed numeric + categorical. Learning resources: missingno library, Kaggle Data Leakage course, Aurélien Géron book Ch.2.",
     "Still pending: architecture parameters. Also need to know what framework/language user will code in (Python + scikit-learn/TensorFlow assumed)."),
    (6, "2026-04-02",
     "Variable mapping & visualization dashboard + data leakage fix",
     "New Stage 2.5 added between outlier treatment and encoding: interactive Streamlit visualization dashboard. Features: (1) Codebook builder — map variable names to question text and variable values to option labels via CSV upload or manual entry. (2) Variable type selection per variable: Categorical(0/1), Ordinal, Numeric — each type has different mapping UI and chart options. (3) Drag-and-drop variable assignment using streamlit-sortables. (4) Chart type selection per variable: Bar/Pie (categorical), Bar/Stacked/Line (ordinal), Histogram/Boxplot/Line (numeric). (5) PDF export of all charts. Charts show post-outlier/missing-value-treated data, before encoding or scaling. Confirmed: train/test split happens BEFORE imputation and scaling — imputer and scaler fitted on train set only, then applied to test set separately.",
     "Still pending: architecture parameters (hidden layers, neurons, activation, loss, optimizer, batch size, epochs). Awaiting framework confirmation (Python/scikit-learn/TensorFlow assumed)."),
    (7, "2026-04-02",
     "Questionnaire doc parsing + tile layout + PyTorch confirmed",
     "Viz dashboard extended: questionnaire document upload (PDF/Word/Excel/TXT) auto-parsed using pdfplumber, python-docx, openpyxl + regex to extract question numbers, text, options, and scale labels into a codebook draft. User edits any misparses inline per tile. Dashboard layout: Grid mode (1–4 tiles/row, user selects) and Slide mode (1–12 tiles/slide, prev/next nav). PyTorch confirmed as NN framework: nn.Module subclass, DataLoader, custom training loop, shap.DeepExplainer for SHAP. Key PyTorch note: nn.CrossEntropyLoss includes softmax internally — do NOT add a separate Softmax layer in the forward pass; output raw logits.",
     "Still pending: hidden layers count, neurons per layer, activation function choice, optimizer, learning rate, batch size, epochs, early stopping."),
    (8, "2026-04-02",
     "Full architecture locked — ALL parameters Done",
     "Architecture confirmed: input_size=5–10 (runtime param, set after feature selection), 2 hidden layers, hidden1=2–4×input_size (16–32), hidden2=hidden1//2 (8–16), funnel shape, ReLU activation, Dropout=0.3 after each hidden layer, no BatchNorm to start. Loss: nn.CrossEntropyLoss (raw logits, no Softmax in forward). Optimizer: Adam(lr=1e-3, weight_decay=1e-4). LR schedule: ReduceLROnPlateau(patience=10, factor=0.5). Batch size: 32. Max epochs: 300 + early stopping (patience=15, monitor val loss, restore best weights). Validation: 5-Fold CV. Init: He uniform (PyTorch default). ALL design decisions are now complete.",
     "READY TO CODE. Suggested order: Stage 0 (Data Audit) → Stage 1 (Missing Values) → Stage 2 (Outliers) → Stage 2.5 (Viz Dashboard) → Stage 3 (Encoding) → Stage 4 (Scaling) → Stages 5–8 (Feature Selection) → NN Training."),
    (9, "2026-04-02",
     "Coding started: Stages 0, 1, 2, 2.5 built + QNR mapping",
     "Code built and tested: Stage 0 (Data Audit — upload, type assignment, missing heatmap, class dist, stats), Stage 1 (Missing values — split-first, imputation), Stage 2 (Outlier detection — IQR/Z-score boxplots, Isolation Forest, treatments), Stage 2.5 (Viz Dashboard — QNR parser, auto column mapper, grid/slide layout, chart selection, PDF export). QNR parser tested on survey.docx: 31 questions parsed. Auto-mapper tested on MyRawdata.xlsx (40 rows, 1171 cols, 493 question cols): 29 questions matched to dataset columns. Segment = target (4 classes). Data observations: A7 has value=8 (Not Enough Experience), A5 uses 1=Aware/2=Not Aware, S8/F1 use 0/1 binary. RowConst/quota/system cols auto-excluded. Google Drive upload script created (upload_to_gdrive.py); zip created for manual upload. All code in app/ folder.",
     "Next: Continue Stages 3–4 (Encoding + Scaling), then Stages 5–8 (Feature Selection), then Stage 9 (PyTorch NN). Also: add option labels to QNR mapping sheet in Excel."),
    (10, "2026-04-03",
     "Full pipeline complete: Stages 3–9 all built. Stage 2.5 codebook editor redesigned.",
     "Stage 2.5 codebook editor fully redesigned: search/filter toolbar, + Add question form, per-question expanders with code/text/type/chart/column editors, st.data_editor options table (SelectboxColumn for dataset_col, dynamic rows), value label quick-editor, delete button. all_cols synced from options table back to tile. Stage 3 (Encoding): auto-detects binary/cat/ordinal cols; One-Hot (pd.get_dummies, test aligned to train), Target Encoding (mean per category, fitted on train), Ordinal Encoding (user-defined order in text area), Label Encoding; per-column override; re-apply support. Stage 4 (Scaling): RobustScaler/StandardScaler/MinMaxScaler; binary cols auto-skipped; per-column override; exclude list; distribution preview. Stage 5 (Correlation): Plotly heatmap (top-50 by variance for large sets), threshold slider, flagged pair table, scatter drill-down, suggested drops, apply/skip. Stage 6 (RF): configurable trees/depth/seed, importance bar chart, cumulative curve, 3 selection modes (Top-N, cumulative %, manual), apply removes cols. Stage 7 (Factor Analysis): Bartlett+KMO, scree plot, loadings table with highlighting, biplot, output as factor scores/high-loaders/both, skip option. Stage 8 (Combination Testing): Forward stepwise + Backward stepwise + Manual explorer tabs, proxy=LogisticRegression, metrics: Accuracy/AUC/McFadden R², results leaderboard, finalize for Stage 9. Stage 9 (Neural Network): TabularNN (2 hidden layers, ReLU, Dropout=0.3, CrossEntropyLoss), Adam+ReduceLROnPlateau, 5-Fold CV, early stopping, final model on full train set, learning curves, confusion matrix, classification report, SHAP via KernelExplainer, McFadden R², re-train option. State management: utils/state.py updated with all 11 stage flags and artefacts. app.py routes all 11 stages.",
     "All stages built. Next: test full pipeline end-to-end with survey.docx + MyRawdata.xlsx. Also: A7 value=8 (NEE) should be treated as missing in Stage 1 — add custom missing value handling. Update NN_Project.zip and re-upload to Google Drive."),
    (11, "2026-04-03",
     "Stage 0 redesigned: sortable column editor table + quick-flag sweep. Column drop panels added to Stages 1 & 2. Stage 2.5 unmatched columns highlighted separately.",
     "Stage 0 — Column editor replaced with a single st.data_editor table (one row per column) showing: Column, Type (SelectboxColumn: numeric/categorical/ordinal), Drop (CheckboxColumn), Missing %, Zero %, Variance, Skewness, Unique, Dtype. Table is sortable by any column header (click to sort). Live summary under table: queued-to-drop count + kept counts by type. Apply changes button writes to session state and reruns immediately. Quick-flag sweep section below uses sliders for missing%/variance/zero% thresholds and 5 buttons (Tick all-null, Tick all-zero, Tick low-var, Tick ALL flagged, Clear all) that pre-tick Drop checkboxes in the editor table via col_editor_state in session state — each button calls st.rerun() for instant reflection. Stage 2.5 codebook editor — unmapped columns now shown in a separate yellow-banner section below matched (green banner) questions. Bulk-exclude button for all unmatched. Matched/unmatched counts shown in summary line above editor. Reusable drop_columns_panel() helper added to utils/ui.py — collapsible expander at top of Stage 1 and Stage 2. Panel shows: per-column stats table (highlighted by flag), 4 quick-add buttons (low-variance/high-missing/high-zero + clear), multiselect with on_change callback, Apply drops now button. All buttons call st.rerun() immediately after state update for real-time reflection. Apply drops now button disabled when nothing selected. Works on df if split not done yet (Stage 1 before split), or X_train/X_test if split done (Stage 2).",
     "Next: end-to-end test with MyRawdata.xlsx. A7 value=8 (NEE) custom missing handling still pending. Re-zip and upload to Google Drive."),
]
for i, row_data in enumerate(session1):
    r = i + 3
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.border    = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill      = fill(LIGHT_BLUE)
    ws2.row_dimensions[r].height = 40

log_widths = [12, 14, 28, 60, 40]
for col, w in enumerate(log_widths, 1):
    set_col_width(ws2, col, w)

# ════════════════════════════════════════════════════════════════
# SHEET 3 – Claude Context Prompt
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Claude Context Prompt")

ws3.merge_cells("A1:B1")
ws3["A1"] = "Copy the text below and paste it as your first message in a new Claude instance"
ws3["A1"].font      = Font(bold=True, color=WHITE, size=12)
ws3["A1"].fill      = fill(DARK_BLUE)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

prompt = """I am building a complete ML pipeline for multiclass classification using PyTorch + Streamlit.
The full pipeline is COMPLETE — all 11 stages are coded and wired. Please read the Session Log
(Sessions 1–10) in the attached Excel for the full decision history.

PROBLEM SUMMARY:
- N = 300–600, 4 output classes, balanced dataset (≤20% variation) → Accuracy is primary metric
- Input: 50–100 raw variables → feature selection pipeline reduces to 5–10 best features
- Mixed input: numeric + categorical

FULL PIPELINE STATUS (all Done):
  Stage 0  — Data Audit          (stage0_audit.py)
  Stage 1  — Missing Values      (stage1_missing.py)
  Stage 2  — Outlier Treatment   (stage2_outliers.py)
  Stage 2.5— Viz Dashboard       (stage2b_viz.py)
  Stage 3  — Encoding            (stage3_encoding.py)
  Stage 4  — Scaling             (stage4_scaling.py)
  Stage 5  — Correlation Filter  (stage5_correlation.py)
  Stage 6  — RF Importance       (stage6_rf.py)
  Stage 7  — Factor Analysis     (stage7_factor.py)
  Stage 8  — Combination Testing (stage8_combtest.py)
  Stage 9  — PyTorch NN          (stage9_nn.py)

ARCHITECTURE (all locked):
  - TabularNN: input_size → hidden1 (2–4×input) → hidden2 (hidden1//2) → 4 logits
  - ReLU, Dropout=0.3, nn.CrossEntropyLoss (NO Softmax in forward — CE handles it)
  - Adam(lr=1e-3, weight_decay=1e-4), ReduceLROnPlateau, 5-Fold CV, early stopping
  - SHAP (KernelExplainer), McFadden pseudo-R², Confusion Matrix, AUC OVR

KEY DATA NOTES:
  - Dataset: MyRawdata.xlsx (40 rows, 1171 cols); target = Segment (4 classes)
  - A7 value=8 = "Not Enough Experience" → treat as missing in Stage 1
  - A5: 1=Aware, 2=Not Aware; S8/F1: 0/1 binary
  - RowConst/quota/system cols auto-excluded in col_mapper.py

CODE LOCATION: C:/Users/azat3/OneDrive/Desktop/NN/
  - app/app.py         — main Streamlit entry point
  - app/stages/        — all stage modules
  - app/utils/         — state.py, ui.py, qnr_parser.py, col_mapper.py
  - requirements.txt   — all dependencies
  - setup_and_run.bat  — double-click to install and launch

TO RUN: streamlit run app/app.py   (from the NN folder)

PENDING TASKS:
  1. End-to-end test: upload survey.docx (Stage 2.5) + MyRawdata.xlsx and walk through all stages
  2. A7 custom missing: add special handling for value=8 in Stage 1 missing value treatment
  3. Re-zip and upload NN_Project.zip to Google Drive for cross-device access
  4. Any further bug fixes or UI improvements found during testing

RECENT UI CHANGES (Session 11):
  - Stage 0: column editor is now a sortable st.data_editor table (Type + Drop + all stats per row)
  - Stage 0: quick-flag sweep buttons pre-tick Drop column; all buttons call st.rerun() immediately
  - Stage 1 & 2: reusable drop_columns_panel() added at top of each stage (collapsible expander)
  - Stage 2.5: matched vs unmatched columns shown in separate colour-coded sections

Please resume by continuing from where we left off — checking for any issues and testing the pipeline."""

ws3["A2"] = "Prompt to paste:"
ws3["A2"].font = bold(size=11)
ws3["A3"] = prompt
ws3["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws3["A3"].fill      = fill(YELLOW)
ws3["A3"].border    = thin_border()
ws3.row_dimensions[3].height = 120
ws3.column_dimensions["A"].width = 100

# ════════════════════════════════════════════════════════════════
# SHEET 4 – QNR Column Mapping
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("QNR Column Mapping")

ws4.merge_cells("A1:G1")
ws4["A1"] = "QNR Question → Dataset Column → Values Mapping  (auto-generated from survey.docx + MyRawdata.xlsx)"
ws4["A1"].font      = Font(bold=True, color=WHITE, size=12)
ws4["A1"].fill      = fill(DARK_BLUE)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

map_headers = ["Q Code", "Variable Type", "Group Type", "Question Text",
               "Dataset Column(s)", "Sample Values", "Scale / Notes"]
for col, h in enumerate(map_headers, 1):
    c = ws4.cell(row=2, column=col, value=h)
    c.font      = bold(WHITE, 10)
    c.fill      = fill(MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
ws4.row_dimensions[2].height = 24

mapping_data = [
    ("S1",  "categorical", "single",    "What is your primary medical specialty?",                            "S1",                               "5",                                     "1 unique value in sample"),
    ("S1A", "scale_7",     "multi_col", "Please indicate your sub-specialty.",                                "S1a, S1aSoftQuota, S1aConst",        "1,2,3",                                 ""),
    ("S2",  "numeric",     "single",    "How many years have you been in practice?",                          "S2",                               "3–40",                                  "Range: 0–99"),
    ("S3",  "ordinal",     "single",    "% professional time in patient care",                                "S3",                               "70,80,85,90,95,98,99,100",              "Range: 0–100%"),
    ("S4",  "ordinal",     "single",    "How many patients did you see in past month?",                       "S4",                               "20–1000+",                              "Range: 0–9999"),
    ("S5",  "grid",        "grid",      "Patients managed in past 12 months by condition",                   "S5_r1_c1, S5_r2_c1",               "r1: THERAPY patients (4–100+); r2: other conditions", "2 rows"),
    ("S5A", "grid",        "grid",      "Patients managed by dependent team members",                        "S5a_r1_c1, S5a_r2_c1",             "r1: THERAPY; r2: other",                "2 rows + RowConst (exclude)"),
    ("S6",  "grid",        "grid",      "% time in each practice setting (sums to 100%)",                    "S6_r3_c1 to S6_r7_c1",             "0–100 per setting",                     "5 active rows; r1,r2 all-null"),
    ("S8",  "multi",       "multi",     "Place of work designation (select all that apply)",                  "S8_1, S8_2, S8_3, S8_4, S8_5",     "0=Not selected, 1=Selected",            "Binary per option"),
    ("S9",  "categorical", "single",    "Family affiliated with pharma/biotech?",                             "S9",                               "2 (1=Yes, 2=No assumed)",               "1 unique value in sample"),
    ("S10A","grid",        "grid",      "Hospital where majority of patients seen",                           "S10a_r1_c1",                       "Hospital IDs (6–110)",                  "Exclude quota/const cols"),
    ("A4",  "open",        "single",    "Maintenance/preventative treatments that come to mind",              "A4",                               "Free text (treatment names)",           "Open-end; 40 unique responses"),
    ("A5",  "grid",        "grid",      "Awareness of maintenance/preventative treatments (per treatment)",   "A5_r1 to A5_r13",                  "1=Aware, 2=Not Aware",                  "13 rows (treatments); _other col for free text"),
    ("A6",  "scale_7",     "multi_col", "Level of knowledge per treatment (1–7)",                             "A6_r1 to A6_r14 (12 active)",      "1=Not at all Knowledgeable … 7=Extremely Knowledgeable", "Only rows where A5=Aware shown"),
    ("A7",  "scale_7",     "multi_col", "Overall satisfaction per treatment (1–7)",                           "A7_r1 to A7_r14 (12 active)",      "1=Not at all Satisfied … 7=Extremely Satisfied; 8=Not Enough Experience", "Value 8 = NEE (treat as missing)"),
    ("A8",  "grid",        "grid",      "New upcoming treatments awareness (open end)",                       "A8_r1_c1 (text), A8_r2_c1 (flag)", "Free text + 0/1 none-aware flag",       ""),
    ("A9",  "grid",        "grid",      "Awareness of new upcoming therapies (per therapy)",                  "A9_r1 to A9_r6",                   "1=Aware, 2=Not Aware",                  "6 rows; r4 cols all-null"),
    ("A10", "scale_7",     "multi_col", "Knowledge of new upcoming therapies (1–7)",                          "A10_r1, r2, r3, r6",               "1=Not at all Knowledgeable … 7=Extremely Knowledgeable", "4 active rows"),
    ("B1",  "grid",        "grid",      "Patient distribution by treatment category (sums to S5_r1)",         "B1_r1_c1 (naive), B1_r2_c1 (on Tx)","0–S5_r1 per row",                      "Must sum to S5_r1"),
    ("B1B", "grid",        "grid",      "Acute treatment received (naive + on-treatment patients)",           "B1bX1_r1_c1, B1bX2_r2_c1",         "0–B1 row values",                       "Exclude RowConst cols"),
    ("B2",  "grid",        "grid",      "Patients on maintenance Tx by category (sums to B1_r2)",             "B2_r1_c1, r2_c1, r3_c1",           "1st/2nd/3rd+ line counts",              "3 rows"),
    ("B3",  "grid",        "grid",      "% patients prescribed each treatment by line (past 12m)",            "B3_r1_c1 to B3_r15 × c1,c2,c3",    "0–200% per cell (combination allowed)", "15 rows × 3 columns = 44 active cols"),
    ("C1",  "scale_7",     "multi_col", "Importance of treatment attributes (1–7)",                           "C1_r1 to C1_r22",                  "1=Not at all Important … 7=Extremely Important", "22 attributes"),
    ("D7A", "grid",        "grid",      "Expected patient distribution next 12m by category",                 "D7a_r1_c2, D7a_r2_c2",             "Must sum to S5_r1",                     "c2 = future; reference col c1 = past"),
    ("D7",  "grid",        "grid",      "% patients prescribed each treatment next 12m",                      "D7_r1_c2 to D7_r17 × c2,c4",        "0–200% per cell",                       "49 active cols across 17 rows × 2 columns"),
    ("E1",  "scale_7",     "multi_col", "Agreement with MOA1 vs MOA2 statements (1–7)",                       "E1_r1, E1_r2",                     "1=Strongly Disagree … 7=Strongly Agree", "2 items"),
    ("E3",  "scale_7",     "multi_col", "Convincingness of MOA1 targeting messages (1–7)",                    "E3_r1 to E3_r5",                   "1=Not at all Convincing … 7=Extremely Convincing", "5 items"),
    ("F1",  "grid",        "grid",      "Awareness of key messages (select all aware)",                       "F1_r1_c1 to F1_r4_c1",             "0=Not aware, 1=Aware",                  "4 messages"),
    ("F2",  "scale_7",     "single",    "Importance of messages for treatment selection (1–7)",               "F2_r1",                            "1=Not at all Important … 7=Extremely Important", "1 item in sample"),
]

type_colors = {
    "categorical": LIGHT_BLUE, "scale_7": YELLOW, "scale_5": YELLOW,
    "grid": GREEN, "multi": ORANGE, "numeric": "D6E4F0",
    "ordinal": "EAD1DC", "open": "F2F2F2", "multi_col": YELLOW,
}

for i, row in enumerate(mapping_data):
    r = i + 3
    row_fill = fill(type_colors.get(row[1], WHITE))
    for col, val in enumerate(row, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.fill      = row_fill
        c.border    = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws4.row_dimensions[r].height = 30

map_widths = [8, 12, 10, 45, 35, 40, 35]
for col, w in enumerate(map_widths, 1):
    set_col_width(ws4, col, w)

ws4.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════
out = r"C:\Users\azat3\OneDrive\Desktop\NN\NN_Design_Tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
