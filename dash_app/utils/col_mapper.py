"""
Auto-maps dataset columns to questionnaire question codes.
Handles:
  - Single columns:   S2, S3, S4  → straight 1:1 mapping
  - Grid columns:     A6_r1_c1, A6_r2_c1 → grouped under A6
  - Multi-select:     S8_1, S8_2, S8_3   → grouped under S8
  - Other suffixes:   _other, _code, _text → attached to parent
"""

import re
import logging
import pandas as pd
from collections import defaultdict

log = logging.getLogger("col_mapper")

# Matches question prefix: S1, A6, B1a, D7a etc.
PREFIX_RE = re.compile(r'^([A-Za-z]\d+[a-z]?)', re.IGNORECASE)

# Strips leading "CODE - " from variable label strings
_CODE_STRIP_RE = re.compile(r'^[A-Za-z]\d+[a-z]?\s*[-–]\s*', re.IGNORECASE)

_CHART_OPTIONS = {
    "categorical": ["Bar chart", "Pie chart", "Donut chart"],
    "ordinal":     ["Bar chart", "Horizontal bar", "Stacked bar", "Line chart"],
    "numeric":     ["Histogram", "Box plot", "Violin plot", "Line chart"],
    "scale_7":     ["Bar chart", "Horizontal bar", "Stacked bar", "Mean line"],
    "scale_5":     ["Bar chart", "Horizontal bar", "Stacked bar", "Mean line"],
    "multi":       ["Bar chart", "Horizontal bar"],
    "grid":        ["Grouped bar", "Heatmap"],
    "open":        ["Word count bar"],
    "single":      ["Bar chart", "Pie chart"],
}


def group_columns(df: pd.DataFrame) -> dict:
    """
    Returns dict:  { 'A6': {'cols': [...], 'group_type': 'grid'|'multi'|'single'},  ... }
    Drops all-null columns automatically.
    """
    # Drop all-null columns
    valid_cols = [c for c in df.columns if df[c].notna().any()]

    groups = defaultdict(list)
    ungrouped = []

    for col in valid_cols:
        m = PREFIX_RE.match(col)
        if m:
            groups[m.group(1).upper()].append(col)
        else:
            ungrouped.append(col)

    result = {}
    for prefix, cols in groups.items():
        # Determine group type
        if len(cols) == 1:
            g_type = "single"
        elif any(re.search(r'_r\d+_c\d+', c) for c in cols):
            g_type = "grid"
        elif all(re.search(r'_\d+$', c) for c in cols):
            g_type = "multi"
        else:
            g_type = "multi_col"

        result[prefix] = {
            "cols":       cols,
            "group_type": g_type,
            "primary":    cols[0],   # main col for simple charts
        }

    # Add ungrouped columns as singles
    for col in ungrouped:
        result[col] = {"cols": [col], "group_type": "single", "primary": col}

    return result


def suggest_var_type(df: pd.DataFrame, cols: list, group_type: str, q_type: str = "") -> str:
    """Suggest variable type from data characteristics."""
    if q_type in ("scale_7", "scale_5"):
        return q_type
    if q_type == "numeric":
        return "numeric"
    if q_type == "open":
        return "open"
    if group_type == "grid":
        return "grid"
    if group_type == "multi":
        return "multi"

    primary = cols[0]
    series = df[primary].dropna()
    if series.empty:
        return "categorical"

    n_unique = series.nunique()
    if pd.api.types.is_float_dtype(series) and n_unique > 10:
        return "numeric"
    if n_unique <= 2:
        return "categorical"
    if n_unique <= 7:
        # Check if values are 1–7 integers (likely scale)
        vals = set(pd.to_numeric(series, errors="coerce").dropna().astype(int).unique())
        if vals <= {1, 2, 3, 4, 5, 6, 7}:
            return "scale_7"
        return "ordinal"
    if n_unique <= 15:
        return "ordinal"
    return "numeric"


def build_codebook(df: pd.DataFrame, parsed_questions: list) -> list:
    """
    Merge parsed questionnaire questions with dataset column groups.
    Returns list of codebook dicts ready for Stage 2.5.
    """
    Q_TYPE_TO_VAR = {
        "scale_7": "scale_7", "scale_5": "scale_5",
        "numeric": "numeric", "open": "open",
        "multi": "multi", "grid": "grid", "single": "categorical",
    }

    col_groups  = group_columns(df)
    q_map       = {q["code"].upper(): q for q in parsed_questions}

    codebook = []
    used_prefixes = set()

    # First pass: questions from QNR that have matching dataset columns
    for q in parsed_questions:
        code = q["code"].upper()
        if code not in col_groups:
            continue
        used_prefixes.add(code)
        grp = col_groups[code]

        var_type = suggest_var_type(
            df, grp["cols"], grp["group_type"],
            q_type=Q_TYPE_TO_VAR.get(q["q_type"], "")
        )

        chart_opts = _CHART_OPTIONS.get(var_type, ["Bar chart"])

        codebook.append({
            "code":         code,
            "question":     q["question"],
            "var_type":     var_type,
            "chart_type":   chart_opts[0],
            "group_type":   grp["group_type"],
            "dataset_col":  grp["primary"],
            "all_cols":     grp["cols"],
            "value_labels": {},
            "scale_low":    q.get("scale_low", ""),
            "scale_high":   q.get("scale_high", ""),
            "scale_points": q.get("scale_points"),
            "pn_notes":     q.get("pn_notes", ""),
            "include":      True,
        })

    # Second pass: dataset column groups with no matching QNR question
    for prefix, grp in col_groups.items():
        if prefix in used_prefixes:
            continue
        # Skip system/metadata cols
        if prefix.lower().startswith(("sys_", "vendor", "account", "panel",
                                       "soft", "hcp", "captcha")):
            continue

        var_type = suggest_var_type(df, grp["cols"], grp["group_type"])
        chart_opts = _CHART_OPTIONS.get(var_type, ["Bar chart"])

        codebook.append({
            "code":         prefix,
            "question":     prefix,   # user can rename
            "var_type":     var_type,
            "chart_type":   chart_opts[0],
            "group_type":   grp["group_type"],
            "dataset_col":  grp["primary"],
            "all_cols":     grp["cols"],
            "value_labels": {},
            "scale_low":    "",
            "scale_high":   "",
            "scale_points": None,
            "pn_notes":     "",
            "include":      False,   # off by default if not in QNR
        })

    return codebook


# ── Excel metadata parsing ─────────────────────────────────────────────────────

def parse_excel_metadata(source) -> dict:
    """
    Read 'Variable Label Information' and 'Value Label Information' sheets
    from a data Excel file (accepts file path or BytesIO).

    Variable Label Information columns:
      A = variable name, B = variable type, E = label text

    Returns:
        {
            "var_labels":   {col_name: label_text},          # col E
            "var_types":    {col_name: type_str},            # col B
            "value_labels": {col_name: {value: label_str}},  # coded columns
            "has_metadata": bool,
        }
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except Exception:
        return {"var_labels": {}, "var_types": {}, "value_labels": {}, "has_metadata": False}

    # Case-insensitive sheet name lookup
    sheet_lower = {s.lower(): s for s in wb.sheetnames}
    var_sheet_key = next(
        (k for k in sheet_lower if "variable" in k and "label" in k), None
    )
    val_sheet_key = next(
        (k for k in sheet_lower if "value" in k and "label" in k), None
    )

    if not var_sheet_key and not val_sheet_key:
        return {"var_labels": {}, "var_types": {}, "value_labels": {}, "has_metadata": False}

    var_labels: dict  = {}
    var_types: dict   = {}
    value_labels: dict = {}

    if var_sheet_key:
        ws = wb[sheet_lower[var_sheet_key]]
        for row in ws.iter_rows(values_only=True, min_row=2):
            if not row or not row[0]:
                continue
            var_name = str(row[0]).strip()
            # col B (index 1) = variable type
            if len(row) > 1 and row[1] is not None:
                var_types[var_name] = str(row[1]).strip()
            # col E (index 4) = label text
            if len(row) > 4 and row[4] is not None:
                var_labels[var_name] = str(row[4]).strip()

    if val_sheet_key:
        ws = wb[sheet_lower[val_sheet_key]]
        current_var = None
        for row in ws.iter_rows(values_only=True, min_row=2):
            if not row:
                continue
            var_name, value, label = row[0], row[1], row[2] if len(row) > 2 else None
            if var_name:
                current_var = str(var_name).strip()
            if current_var and value is not None and label is not None:
                value_labels.setdefault(current_var, {})[value] = str(label).strip()

    return {
        "var_labels":   var_labels,
        "var_types":    var_types,
        "value_labels": value_labels,
        "has_metadata": True,
    }


def _vl_sort_key(item):
    v = item[0]
    if isinstance(v, (int, float)):
        return (0, float(v))
    return (1, str(v))


def _col_b_to_var_type(raw: str) -> str:
    """Map raw col-B type string to one of our canonical var-type keys."""
    if not raw:
        return ""
    s = raw.lower().strip()
    if "numeric" in s or "integer" in s or "float" in s or "double" in s:
        return "numeric"
    if "string" in s or "char" in s or "text" in s or "open" in s:
        return "open"
    if "ordinal" in s:
        return "ordinal"
    if "scale" in s:
        if "7" in s:
            return "scale_7"
        if "5" in s:
            return "scale_5"
        return "ordinal"
    if "multi" in s or "multiple" in s:
        return "multi"
    if "grid" in s or "matrix" in s:
        return "grid"
    if "categ" in s or "nominal" in s or "single" in s:
        return "categorical"
    return ""


def _norm_text(s: str) -> str:
    """Normalize option / label text for comparison."""
    if not s:
        return ""
    s = _CODE_STRIP_RE.sub("", str(s))
    # remove non-alphanum aside from spaces; collapse whitespace; lowercase
    s = re.sub(r"[\s_]+", " ", s.lower())
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _overlap_score(a_norm: str, b_norm: str) -> float:
    """
    Return a loose overlap score between two normalized strings.
    1.0 if substring containment; otherwise fraction of shared words
    over the smaller bag of words.
    """
    if not a_norm or not b_norm:
        return 0.0
    if a_norm == b_norm:
        return 1.0
    # Word-boundary containment: require whole-token match to avoid
    # "p1" matching inside "p10", "p11", etc.
    wa = set(a_norm.split())
    wb = set(b_norm.split())
    if not wa or not wb:
        return 0.0
    if wa == wb:
        return 1.0
    # Containment only when one side is entirely made of whole words of the other
    if wa <= wb or wb <= wa:
        return 0.95
    return len(wa & wb) / min(len(wa), len(wb))


def _best_qnr_match(ce_norm: str, qnr_norms: list, qnr_opts: list,
                    threshold: float = 0.4):
    """
    Return (best_qnr_opt, best_score) or (None, 0.0).
    `qnr_norms` is pre-normalized QNR options list (same length as qnr_opts).
    """
    if not ce_norm or not qnr_opts:
        return None, 0.0
    best_opt, best_score = None, 0.0
    for opt, norm_o in zip(qnr_opts, qnr_norms):
        score = _overlap_score(ce_norm, norm_o)
        # Prefer longer/more specific match on equal score
        if score > best_score or (score == best_score and best_opt is not None and len(norm_o) > len(_norm_text(best_opt))):
            best_score, best_opt = score, opt
    if best_score >= threshold:
        return best_opt, best_score
    return None, best_score


def _infer_col_headers_from_suffixes(code: str, cols: list, n_opts: int) -> list:
    """
    When QNR `table_col_headers` is empty but we can detect a grid-like
    ratio, try to derive column headers from the dataset column suffixes.

    Strategy: strip the question prefix from each col, then find the
    *second* varying token (the first token usually indexes the option).
    Returns the ordered unique inferred headers, or [] if none found.
    """
    if not cols or n_opts <= 0:
        return []
    if len(cols) % n_opts != 0:
        return []
    n_headers = len(cols) // n_opts
    if n_headers <= 1:
        return []

    # Strip the question prefix from each column
    prefix_re = re.compile(
        r"^" + re.escape(code) + r"[_-]?", re.IGNORECASE
    )
    suffixes = [prefix_re.sub("", c) for c in cols]

    # Split on _ or -
    parts = [re.split(r"[_\-]", s) for s in suffixes]

    # Pick the first token position that has exactly n_headers unique values
    for pos in range(max((len(p) for p in parts), default=0)):
        seen = []
        for p in parts:
            if pos < len(p):
                tok = p[pos]
                if tok not in seen:
                    seen.append(tok)
        if len(seen) == n_headers:
            return seen
    return []


def log_assignments(questions_enriched: list, col_groups: dict) -> None:
    """
    Emit a structured DEBUG-level log of how each question was resolved.
    Each enriched question must carry `_mapping_source` and `col_assignments`
    (annotations attached by merge_qnr_with_metadata).
    """
    for q in questions_enriched:
        code   = q.get("code", "?")
        q_type = q.get("q_type", "")
        source = q.get("_mapping_source", "n/a")
        warns  = q.get("_mapping_warnings", [])
        assigns = q.get("col_assignments", {}) or {}

        log.debug(
            "── %s  q_type=%s  source=%s  n_opts=%d  n_cols=%d",
            code, q_type, source,
            len(q.get("options", [])),
            len((col_groups.get(code) or {}).get("cols", [])),
        )
        for w in warns:
            log.debug("   ! warning: %s", w)

        # Per-column assignment line (truncated)
        grp = col_groups.get(code) or {}
        for col in grp.get("cols", []):
            assigned = assigns.get(col)
            if assigned:
                shown = ", ".join(str(a)[:60] for a in assigned[:3])
                if len(assigned) > 3:
                    shown += f", … (+{len(assigned) - 3})"
                log.debug("   %s → [%s]", col, shown)
            else:
                log.debug("   %s → UNASSIGNED", col)


def merge_qnr_with_metadata(questions: list, meta: dict, col_groups: dict) -> list:
    """
    Enrich QNR-parsed questions with per-column option assignments using
    Excel metadata (Variable Label Information sheet).

    Logic per question:
      1. Collect all dataset columns belonging to this question code.
      2. Pull col-E (label) and col-B (type) text for each column.
      3. Determine the canonical option set:
           · If ≥1 overlap between QNR options and col-E labels → INTERSECTION
             (keep only QNR options that match ≥1 col-E label; QNR text is
             the canonical string).
           · Else → QNR options only (QNR is ground truth).
           · If no QNR options at all → use unique cleaned col-E labels.
      4. Grid expansion: when n_cols == n_options * (n_extra_table_cols),
         expand each QNR option as "{col_header} {opt}" and assign
         sequentially. If table_col_headers is empty, try to infer headers
         from the column name suffixes.
      5. Overassignment safety: any column still without an assignment gets
         a positional-suffix match or, as last resort, all final options.
      6. Meta var type is mapped from col-B via _col_b_to_var_type.

    Every dataset column belonging to a known question ends up with a
    non-empty entry in q["col_assignments"].
    """
    var_labels = meta.get("var_labels", {}) if meta else {}
    var_types  = meta.get("var_types",  {}) if meta else {}

    enriched: list = []

    for q in questions:
        code     = q["code"].upper()
        grp      = col_groups.get(code)
        qnr_opts = list(q.get("options", []))

        if not grp:
            enriched.append(q)
            continue

        cols: list = list(grp["cols"])
        warnings: list = []

        # ── Step 1: gather col-E labels per column ───────────────────────────
        col_e_raw: dict = {}          # col → raw label from col E
        col_e_clean: dict = {}        # col → "CODE - "-stripped label
        for col in cols:
            raw = var_labels.get(col, "")
            if raw:
                col_e_raw[col] = raw
                col_e_clean[col] = _CODE_STRIP_RE.sub("", raw).strip()

        qnr_norms   = [_norm_text(o) for o in qnr_opts]
        ce_pairs    = [(c, col_e_clean[c]) for c in cols if col_e_clean.get(c)]
        ce_norms    = [_norm_text(t) for _, t in ce_pairs]

        # ── Step 2: determine whether QNR ↔ col-E sets overlap ───────────────
        any_overlap = False
        if qnr_opts and ce_pairs:
            for cen in ce_norms:
                for qn_ in qnr_norms:
                    if _overlap_score(cen, qn_) >= 0.4:
                        any_overlap = True
                        break
                if any_overlap:
                    break

        # ── Step 3: choose the canonical option set ──────────────────────────
        # Strategy controls the option pool; canonical text is ALWAYS the QNR
        # wording when available (more descriptive than col-E).
        used_qnr_idx: set = set()       # QNR options that matched something
        col_to_qnr: dict  = {}          # col → QNR option text (via col-E match)

        if qnr_opts and ce_pairs and any_overlap:
            # Intersection: for each col, find best QNR option via col-E
            for (col, ce_text), ce_norm in zip(ce_pairs, ce_norms):
                m_opt, score = _best_qnr_match(ce_norm, qnr_norms, qnr_opts)
                if m_opt is not None:
                    col_to_qnr[col] = m_opt
                    used_qnr_idx.add(qnr_opts.index(m_opt))
            # canonical option set = QNR options that had at least one col-E match,
            # order preserved from QNR
            intersection_opts = [
                o for i, o in enumerate(qnr_opts) if i in used_qnr_idx
            ]
            # If intersection drops too many options (less than a third),
            # fall back to full QNR list so nothing disappears from the UI.
            if len(intersection_opts) >= max(1, len(qnr_opts) // 3):
                final_opts = intersection_opts
                source = "intersection"
                if len(intersection_opts) < len(qnr_opts):
                    warnings.append(
                        f"intersection kept {len(intersection_opts)}/{len(qnr_opts)} "
                        "QNR options"
                    )
            else:
                final_opts = list(qnr_opts)
                source = "qnr_only"
                warnings.append(
                    f"intersection kept only {len(intersection_opts)}/{len(qnr_opts)} "
                    "opts — falling back to full QNR option set"
                )
        elif qnr_opts:
            final_opts = list(qnr_opts)
            source = "qnr_only"
            if ce_pairs and not any_overlap:
                warnings.append(
                    "col_e labels found but no overlap with QNR — using QNR only"
                )
        elif ce_pairs:
            # No QNR options; derive unique cleaned col-E labels
            seen, final_opts = set(), []
            for _, ce_text in ce_pairs:
                if ce_text and ce_text not in seen:
                    seen.add(ce_text)
                    final_opts.append(ce_text)
            source = "col_e_only"
        else:
            final_opts = []
            source = "empty"

        # ── Step 4: grid expansion detection ─────────────────────────────────
        n_vars      = len(cols)
        n_opts      = len(final_opts)
        table_n_cols = int(q.get("table_n_cols", 1) or 1)
        headers_raw  = list(q.get("table_col_headers", []) or [])
        # Strip empty column headers and the first (variable name) column header
        col_headers = [h for h in headers_raw if h and h.strip()]

        expanded_opts: list = []
        grid_expanded = False

        if n_opts > 0 and n_vars > 0:
            # Case A: QNR gave us table_col_headers and arithmetic matches
            if col_headers and n_vars == n_opts * len(col_headers):
                expanded_opts = [
                    f"{ch.strip()} {opt}".strip()
                    for opt in final_opts
                    for ch in col_headers
                ]
                grid_expanded = True
            # Case B: no headers but n_vars is a clean multiple of n_opts → infer
            elif (not col_headers
                  and n_vars > n_opts
                  and n_vars % n_opts == 0):
                inferred = _infer_col_headers_from_suffixes(
                    code, cols, n_opts
                )
                if inferred:
                    col_headers = inferred
                    expanded_opts = [
                        f"{ch.strip()} {opt}".strip()
                        for opt in final_opts
                        for ch in col_headers
                    ]
                    grid_expanded = True
                    warnings.append(
                        f"inferred {len(inferred)} col headers from suffixes: "
                        f"{inferred}"
                    )
            # Case C: legacy pattern using table_n_cols
            elif (table_n_cols > 2
                  and col_headers
                  and n_vars == n_opts * (table_n_cols - 1)):
                # only the non-variable columns contribute headers
                hdrs = col_headers[-(table_n_cols - 1):]
                expanded_opts = [
                    f"{ch.strip()} {opt}".strip()
                    for opt in final_opts
                    for ch in hdrs
                ]
                grid_expanded = True

        # ── Step 4b: cross-validate option count vs dataset column count ────────
        # If QNR gave fewer options than dataset columns imply, the QNR parse
        # is likely incomplete (e.g. table rows were silently dropped).
        if n_opts > 0 and n_vars > 0 and not grid_expanded:
            n_col_hdrs = len(col_headers) if col_headers else 1
            expected_opts = n_vars // n_col_hdrs
            if n_opts < expected_opts * 0.7 and expected_opts > n_opts:
                msg = (
                    f"{code}: QNR has {n_opts} option(s) but {n_vars} dataset "
                    f"column(s) suggest ~{expected_opts} — QNR parsing may be "
                    f"incomplete (check table structure in the source document)"
                )
                warnings.append(msg)
                log.warning(msg)

        # ── Step 5: build col_assignments for every column ───────────────────
        col_assignments: dict = {}

        if grid_expanded and expanded_opts:
            # Sequential assignment of expanded options
            for i, col in enumerate(cols):
                if i < len(expanded_opts):
                    col_assignments[col] = [expanded_opts[i]]
            # Use the expanded set as the question's option set so the UI
            # shows them as separate chips
            final_opts = expanded_opts
            source = "grid_expansion"
        else:
            # Non-grid: first try col-E → QNR mapping
            for col in cols:
                if col in col_to_qnr:
                    col_assignments[col] = [col_to_qnr[col]]

            # Remaining columns with col-E but no QNR match → use cleaned label
            for col in cols:
                if col in col_assignments:
                    continue
                ce = col_e_clean.get(col, "")
                if ce and source in ("col_e_only",):
                    col_assignments[col] = [ce]

        # ── Step 6: overassignment safety — no column left blank ─────────────
        for col in cols:
            if col_assignments.get(col):
                continue
            # positional suffix: _1 → opts[0], _2 → opts[1] …
            placed = False
            m_num = re.search(r"[_\-](\d+)(?:[_\-]\d+)?$", col)
            m_let = re.search(r"[A-Z]\d+([a-z])$", col, re.IGNORECASE)
            if m_num and final_opts:
                idx = int(m_num.group(1)) - 1
                if 0 <= idx < len(final_opts):
                    col_assignments[col] = [final_opts[idx]]
                    placed = True
            if not placed and m_let and final_opts:
                idx = ord(m_let.group(1).lower()) - ord("a")
                if 0 <= idx < len(final_opts):
                    col_assignments[col] = [final_opts[idx]]
                    placed = True
            if not placed:
                # overassign all options — better than leaving blank
                if final_opts:
                    col_assignments[col] = list(final_opts)
                else:
                    col_assignments[col] = []
                warnings.append(
                    f"{col}: no option match — overassigned "
                    f"{len(final_opts)} opts as fallback"
                )

        # ── Step 7: build enriched question ──────────────────────────────────
        q_new = dict(q)
        q_new["options"]            = final_opts
        q_new["col_assignments"]    = col_assignments
        q_new["_mapping_source"]    = source
        q_new["_mapping_warnings"]  = warnings
        if grid_expanded:
            q_new["table_col_headers"] = col_headers

        # Var type from col B (primary column)
        primary_type_raw = var_types.get(cols[0], "") if cols else ""
        mapped_type      = _col_b_to_var_type(primary_type_raw)
        if mapped_type:
            q_new["meta_var_type"] = mapped_type

        enriched.append(q_new)

    # Debug log for the whole batch
    try:
        log_assignments(enriched, col_groups)
    except Exception as e:  # never break the pipeline on logging
        log.debug("log_assignments failed: %s", e)

    return enriched


def build_questions_from_metadata(var_labels: dict, value_labels: dict,
                                   col_groups: dict) -> list:
    """
    Build question dicts (same format as qnr_parser output) from Excel metadata.

    Logic per group type:
      single   → options come from value_labels of the primary column (sorted by value)
      multi    → options come from stripped var_labels of each column (1 chip per col)
      grid     → same as multi
      multi_col → same as multi
    """
    questions = []

    for prefix, grp in col_groups.items():
        cols       = grp["cols"]
        primary    = grp["primary"]
        group_type = grp["group_type"]

        # ── Question text ─────────────────────────────────────────────────────
        raw_label = var_labels.get(primary, "")
        q_text    = _CODE_STRIP_RE.sub("", raw_label).strip() or prefix

        # ── Options ───────────────────────────────────────────────────────────
        if group_type == "single":
            vl      = value_labels.get(primary, {})
            options = [lbl for _, lbl in sorted(vl.items(), key=_vl_sort_key)]
        else:
            # Multi-column: one chip per column from the column's own label
            seen    = set()
            options = []
            for col in cols:
                lbl      = var_labels.get(col, "")
                stripped = _CODE_STRIP_RE.sub("", lbl).strip()
                if stripped and stripped not in seen:
                    options.append(stripped)
                    seen.add(stripped)

        # ── Type inference ────────────────────────────────────────────────────
        if group_type == "grid":
            q_type = "grid"
        elif group_type in ("multi", "multi_col"):
            q_type = "multi"
        elif not options:
            q_type = "numeric"
        else:
            vl   = value_labels.get(primary, {})
            vals = set(vl.keys())
            n    = len(options)
            if n <= 2:
                q_type = "single"
            elif n <= 7 and vals and all(isinstance(v, (int, float)) for v in vals):
                int_vals = {int(v) for v in vals}
                if int_vals <= {1, 2, 3, 4, 5, 6, 7}:
                    q_type = f"scale_{n}"
                else:
                    q_type = "ordinal"
            elif n <= 15:
                q_type = "ordinal"
            else:
                q_type = "categorical"

        questions.append({
            "code":         prefix,
            "full_code":    prefix,
            "question":     q_text,
            "q_type":       q_type,
            "options":      options,
            "pn_notes":     "",
            "scale_low":    "",
            "scale_high":   "",
            "scale_points": None,
        })

    return questions
